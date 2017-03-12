from openpyxl import load_workbook
import MySQLdb
import time
import sys
reload(sys)
sys.setdefaultencoding('utf8')

file = 'Copy-of-final_company_updated.xlsx'
workbook_name = 'Sheet1'


def connect_to_db():
    return MySQLdb.connect("localhost", "root", "Jindal.house36", "placecom_feed_development")


def enable_unicode_on_db(db, cursor):
    db.set_character_set('utf8')
    cursor.execute('SET NAMES utf8;')
    cursor.execute('SET CHARACTER SET utf8;')
    cursor.execute('SET character_set_connection=utf8;')

def push_to_mysql(data):
    db = connect_to_db()
    cursor = db.cursor()
    enable_unicode_on_db(db, cursor)

    for company in data:
        stmt = "INSERT INTO companies (id, compname ,created_at, updated_at) VALUES(NULL, '%s','%s','%s' )" % (company ,time.strftime('%Y-%m-%d %H:%M:%S'),time.strftime('%Y-%m-%d %H:%M:%S'))
        print stmt
        cursor.execute(stmt)

    db.commit()
    db.close()


def read_data(file, workbook_name):
    header = 0
    wb = load_workbook(file)
    ws = wb[workbook_name]
    data = []
    for row in ws.iter_rows():
        if header == 0:
            header = 1
            continue

        data.append(row[1].value)

    data_set = set(data)
    data = []
    for company in data_set:
        data.append(company)

    push_to_mysql(data)

read_data(file, workbook_name)