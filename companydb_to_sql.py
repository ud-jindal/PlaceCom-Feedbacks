from openpyxl import load_workbook
import MySQLdb
import time
import sys
reload(sys)
sys.setdefaultencoding('utf8')
file = 'Copy-of-final_company_updated.xlsx'
workbook_name = 'Sheet1'


def connect_to_db():
    return MySQLdb.connect("localhost", "root", "Jindal.house36", "placecom_feed_production")


def enable_unicode_on_db(db, cursor):
    db.set_character_set('utf8')
    cursor.execute('SET NAMES utf8;')
    cursor.execute('SET CHARACTER SET utf8;')
    cursor.execute('SET character_set_connection=utf8;')


def push_to_mysql(data):
    db = connect_to_db()
    cursor = db.cursor()
    enable_unicode_on_db(db, cursor)

    # stmt = "INSERT INTO company_data (id, name, contact_name, email, phone_no, created_at, updated_at) VALUES(NULL, '%s', '%s', '%s', '%s', '%s','%s' )" %\
    #        (str(data[1]).encode('utf8'), str(data[2]).encode('utf8') ,
    #         str(data[4]).encode('utf8'), str(data[5]).encode('utf8'),time.strftime('%Y-%m-%d %H:%M:%S'),time.strftime('%Y-%m-%d %H:%M:%S') )
    stmt = "INSERT INTO company_data (id, name, contact_name, email, phone_no, created_at, updated_at) VALUES(NULL, '%s', '%s', '%s', '%s', '%s','%s' )" %\
           (data[1].encode('utf8'), data[2].encode('utf8')+ " "+data[3].encode('utf8')  ,
            data[4].encode('utf8'), data[5],time.strftime('%Y-%m-%d %H:%M:%S'),time.strftime('%Y-%m-%d %H:%M:%S') )
   
    print stmt
    cursor.execute(stmt)
    db.commit()


def read_data(file, workbook_name):
    header = 0
    wb = load_workbook(file)
    ws = wb[workbook_name]
    for row in ws.iter_rows():
        data = []
        if header == 0:
            header = 1
            continue

        for cell in row:
            if cell.value is not None:
                data.append(cell.value)
            else:
                data.append("")
        push_to_mysql(data)


read_data(file, workbook_name)
